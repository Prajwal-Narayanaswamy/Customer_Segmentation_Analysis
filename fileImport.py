import csv
from openpyxl import load_workbook, Workbook

from importer import *


class FileImporter:

    def __init__(self, filePath, defaultHeaders=None, defaultDataTypes=None, rowDataType=ROW_TYPE_DICT,
                 noneStrings=DEFAULT_NONE_STRINGS, rowLimit=None):
        """A class used to read, scrub, analyze, and write data from/to Excel and CSV

        :param str filePath: Name of the file to read
        :param dict defaultHeaders: A key/value pair of header in raw file and the new name of header
        :param dict defaultDataTypes: A key/value pair of header and the function or data type to convert each value to
        :param str rowDataType: A string indicating the data type that each row should be processed as
        """
        self.defaultHeaders = defaultHeaders
        self.defaultDataTypes = defaultDataTypes
        self.rowDataType = rowDataType
        self.noneStrings = noneStrings

        self.fileReader = self.getFileReader(filePath)
        self.headers = self.processHeaders(list(next(self.fileReader)))
        print(f'Headers: {self.headers}')
        self.data = self.getData(rowLimit)

    ##### READ DATA #########
    def printRows(self, numRows):
        for idx, row in enumerate(self.data):
            print(row)
            if idx == numRows - 1:
                break

    def getData(self, rowLimit):
        """ Get all row data from the file import and return a processed list of records
        :return: list
        """
        data = []
        for idx, row in enumerate(self.fileReader):
            if rowLimit and idx == rowLimit:
                break
            processedRow = self.processRow(row)
            if idx == 0:
                headersAndVals = zip(self.headers, processedRow) if self.rowDataType != ROW_TYPE_DICT else processedRow.items()
                for header, val in headersAndVals:
                    print(f'{header} type is {type(val)}')
            data.append(processedRow)

        # CSV file must be closed manually
        if self.fileType == FILE_TYPE_CSV:
            self.file.close()

        return data

    def processRow(self, row):
        """ Process the row into the data type specified by FileImporter
        :return: list|tuple|dict
        """
        processedRow = []
        for header, val in zip(self.headers, row):
            defaultDataType = self.defaultDataTypes.get(header) if self.defaultDataTypes else None
            processedRow.append(self.processValue(val, defaultDataType))
        if self.rowDataType == ROW_TYPE_LIST:
            return processedRow
        if self.rowDataType == ROW_TYPE_DICT:
            return {header: val for header, val in zip(self.headers, processedRow)}
        if self.rowDataType == ROW_TYPE_TUPLE:
            return tuple(processedRow)

    def processValue(self, val, defaultDataType):
        """ Get a processed value. Handles None, null strings, and function conversions
        :param func|class defaultDataType: A function or class which converts the value into the appropriate format
        """
        if val is None:
            return val
        if isinstance(val, str):
            val = val.strip()
            if val.lower() in self.noneStrings:
                return None
        if defaultDataType:
            val = defaultDataType(val)
        return val

    def getFileReader(self, filePath):
        """ Get an iterator used to get each row in a CSV or XLSX file
        """
        if f'.{FILE_TYPE_CSV}' in filePath:
            self.file = open(filePath)
            self.fileType = FILE_TYPE_CSV
            return csv.reader(self.file)
        elif f'.{FILE_TYPE_XLS}' in filePath:
            self.file = load_workbook(filePath, data_only=True)
            worksheet = self.file.active
            self.fileType = FILE_TYPE_XLS
            return worksheet.iter_rows(values_only=True)
        else:
            raise(ValueError('Unsupported file type'))

    def processHeaders(self, headers):
        """ Convert headers to formatted names
        """
        if not self.defaultHeaders:
            return headers
        newHeaders = []
        for idx, header in enumerate(headers):
            headerByIdx = self.defaultHeaders.get(idx)
            headerByVal = self.defaultHeaders.get(header)
            if headerByIdx:
                newHeaders.append(headerByIdx)
            elif headerByVal:
                newHeaders.append(headerByVal)
            else:
                newHeaders.append(header)
        return newHeaders

    ##### WRITE DATA #########
    def writeCsvFile(self, fileName, data=None, headers=None):
        """ Write data to a new CSV file
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list data: If provided, will be used instead of the FileImporter's internal data property
        :param list headers: If provided, will be used instead of the FileImporter's internal header property
        """
        dataToWrite = data or self.data
        headersToWrite = headers or self.headers

        with open(f'{DATA_FILE_OUTPUT_PATH}{fileName}.csv', 'w') as csvFile:
            csvWriter = csv.writer(csvFile)
            csvWriter.writerow(headersToWrite)
            for row in dataToWrite:
                csvWriter.writerow(self.formatRow(row))

    def writeExcelFile(self, fileName, sheetsConfig=None):
        """ Write data to a new XLSX file and return the workbook
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list sheetsConfig: [{'title': , 'data': , 'headers':},...] If provided, will be used instead
        of the FileImporter's internal data property
        """
        if not sheetsConfig:
            return

        workbook = Workbook()

        for idx, config in enumerate(sheetsConfig):
            # Create new worksheet
            if idx == 0:
                worksheet = workbook.active
            else:
                worksheet = workbook.create_sheet()
            worksheet.title = config.get('title', None)

            # Add data to sheet
            dataToWrite = config.get('data', None) or self.data
            headersToWrite = config.get('headers', None) or self.headers
            worksheet.append(headersToWrite)
            for row in dataToWrite:
                worksheet.append(self.formatRow(row))

        # Save and return workbook
        workbook.save(f'{DATA_FILE_OUTPUT_PATH}{fileName}.xlsx')
        return workbook

    def formatRow(self, row):
        """ Process row to make sure it is an appropriate data format to write to a CSV or XSLX file.
        """
        if isinstance(row, (tuple, list)):
            return row
        if isinstance(row, dict):
            return list(row.values())
        raise ValueError('Must use a row data type of tuple, list, or dict')

    ##### ANALYZE DATA #########
    # [(customerKey,), (sex, isMarried)]
    # {('customerKey'): {'a-11243d': [<record1>, <record2>]}, (sex, isMarried): {}}
    def getGroupData(self, groupings):
        dataGroups = {group: {} for group in groupings}

        for group in groupings:
            currentGroup = dataGroups[group]
            for record in self.data:
                groupKey = self.getGroupKey(record, group)
                if groupKey in currentGroup:
                    currentGroup[groupKey].append(record)
                else:
                    currentGroup[groupKey] = [record]

        return dataGroups

    def setGroupData(self, groupings):
        self.dataGroups = self.getGroupData(groupings)

    def getGroupKey(self, record, group):
        if self.rowDataType != ROW_TYPE_DICT:
            record = {header: val for header, val in zip(self.headers, record)}

        return tuple(record[column] for column in group)
